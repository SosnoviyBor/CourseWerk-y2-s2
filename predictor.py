import pickle
import pandas as pd
import openpyxl as excel
import os

from sklearn.preprocessing import PolynomialFeatures, StandardScaler
from sklearn.linear_model import LinearRegression
from sklearn.pipeline import Pipeline

def __create_dummies (data:dict, key:str, dummies:list) -> dict:
	"""
	data = dictionary to pump
	key = key, which value is to be transformed
	dummies = list of dummy keys to create
	"""
	modified = False
	# dummy duplication machine!
	for dummie in dummies:
		if data[key].lower() == f"{dummie}".lower():
			data.update({f"{dummie}":1})
			modified = True
		else: 
			data.update({f"{dummie}":0})
	# data wasn't modified
	if not modified:
		raise ValueError(f"This {dummie} is not supported currently. Or you misspelled it")
	# drop the excess key
	data.pop(key)
	return data

def predict(data:dict|pd.Series, model_name:str, to_excel:bool=False) -> list|pd.Series:
	"""
	If data is dictionary, it should look like this (any order):\n
	data = {
		"firm": str,
		"year": int,
		"price": int,
		"mileage": int,
		"fuelType": str,
		"transmission": str,
		"kmpl": int|float,
		"engineSize": int|float
	}
	"""
	# generate init dataframe if data is dictionary
	if type(data) == type({}):
		# create all dummies
		data = __create_dummies(data, "firm", [
			"audi", "bmw", "ford", "hyundai", "mercedes",
			"skoda", "toyota", "vauxhall", "volkswagen"
		])
		data = __create_dummies(data, "transmission", ['Manual', 'Automatic', 'Semi-Auto', 'Other'])
		data = __create_dummies(data, "fuelType", ['Petrol', 'Diesel', 'Hybrid', 'Other', 'Electric'])
		# insert all values in lists so that data could be transformed to dataframe
		for key in data.keys():
			data[key] = [data[key]]
		# moment that we all were waiting for!
		df = pd.DataFrame.from_dict(data)
		# rearrange dataframe columns order
		df = df[["audi","bmw","ford","hyundai","mercedes",			# firms 1
				 "skoda","toyota","vauxhall","volkswagen",			# firms 2
				 "year","mileage","kmpl","engineSize",				# default data
				 "Automatic","Manual","Other","Semi-Auto",			# transmission
				 "Diesel","Electric","Hybrid","Other","Petrol"]]	# fuel type
	else:
		df = data.drop("price", axis=1)

	# get the model
	with open(f"models/{model_name}", "rb") as file:
		model = pickle.load(file)
		# do magic bullshittery
		rawResult = model.predict(df)
		# format result to same format
		result = []
		for cost in rawResult:
			# multi returns array but poly returns arrya of arrays of 1 value
			# WTFFFFFF
			if type(cost) == type("<class 'numpy.ndarray'>"):
				cost = cost[0]
			if cost < 0:
				cost = abs(cost)
			cost = int(cost)
			result.append(cost)

		if to_excel:
			# write result to excel file
			wb = excel.load_workbook("test results.xlsx")
			ws = wb.create_sheet(model_name)
			ws["A1"] = model_name
			ws["A2"] = "actual"
			ws["B2"] = "predicted"
			ws["C2"] = "diffference"
			ws["D2"] = "diffference%"

			for i in range(len(result)):
				actual = int(data.iloc[i]["price"])
				predicted = int(result[i])
				
				# no clue how this thing comes up with bazillions but ok
				if predicted < 10000000:
					ws[f"A{i+3}"] = actual
					ws[f"B{i+3}"] = predicted
					ws[f"C{i+3}"] = f"=ABS(A{i+3}-B{i+3})"
					ws[f"D{i+3}"] = f"=(C{i+3}/A{i+3}*100)"
				else: ws[f"A{i+3}"] = "DELETE THIS ROW"
			
			# writing cool statistics stuff
			ws["F2"] = "Avg diff%"
			ws["F3"] = f"=AVERAGE(D3:D{3+len(result)})"
			ws["G2"] = "MSE"
			ws["G3"] = f"=SQRT(SUMSQ(C3:C{3+len(result)})/COUNT(C3:C{3+len(result)}))"
			
			wb.save(filename="test results.xlsx")
			print(f"Model {model_name} finished calculating!")
			return
		else:
			return result

def train(df:pd.Series, regression:str, length:int, save:bool, degree:str|int="",):
	df = df.copy(deep=True)
	trainingData = df.drop("price", axis=1)
	
	match regression:
		case "multi":
			# le small math
			pipa = LinearRegression.fit(LinearRegression(),X=trainingData[:length], y=df["price"][:length])
			
			# saving model for future usage or returning for whatever reason
			if save:
				filepath = f"models/{length} {regression}"
				if "bmw" not in trainingData.columns:
					filepath += " nof"
				with open(filepath, "wb") as file:
					pickle.dump(pipa, file)
			else:
				return pipa
		case "poly":
			# le big math
			blueprint = [('scale',StandardScaler()),('polynomial',PolynomialFeatures(degree=degree)),('mode',LinearRegression())]
			pipa = Pipeline(blueprint)
			pipa.fit(trainingData[:length], df[["price"]][:length])
			
			# saving model for future usage or returning for whatever reason
			if save:
				filepath = f"models/{length} {regression} {degree}"
				if "bmw" not in trainingData.columns:
					filepath += " nof"
				with open(filepath, "wb") as file:
					pickle.dump(pipa, file)
			else:
				return pipa
		case _:
			raise ValueError("Expected regression to be 'multi' or 'poly'")
	
	# its literally the first time i actually used ternary operator seriously
	is_nof = "" if "bmw" in df.columns else " nof"
	print(f"Model '{length} {regression} {degree}{is_nof}' successfully trained!")
	return

def train_and_predict (instructions:list, regression:str, save_model:bool, do_tests:bool, to_excel:bool=False):
	"""
	instructions = [
		[dataframe, length, degree (if regression == "poly")],\n
		...
	]
	"""
	match regression:
		case "multi":
			for args in instructions:
				# check for firms
				if not "bmw" in args[0].columns:
					filename = f"{args[1]} multi nof"
				else:
					filename = f"{args[1]} multi"
				
				# check if this model is aleady available
				if not os.path.exists(f"models/{filename}"):
					train(args[0][:args[1]], "multi", args[1], save_model)
				
				# maybe you wanted only to train your models?
				if do_tests:
					return predict(args[0][args[1]:], filename, to_excel)
		case "poly":
			for args in instructions:
				# check for firms
				if not "bmw" in args[0].columns:
					filename = f"{args[1]} poly {args[2]} nof"
				else:
					filename = f"{args[1]} poly {args[2]}"
				
				# check if this model is aleady available
				if not os.path.exists(f"models/{filename}"):
					train(args[0][:args[1]], "poly", args[1], save_model, degree=args[2])
				
				# maybe you wanted only to train your models?
				if do_tests:
					return predict(args[0][args[1]:], filename, to_excel)